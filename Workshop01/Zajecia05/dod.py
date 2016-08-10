'''
Created on 9 sie 2016

@author: giberkrz
'''
from openpyxl import Workbook
from openpyxl.writer.write_only import WriteOnlyCell
from openpyxl.comments import Comment
from openpyxl.styles import Font
from openpyxl import load_workbook
from openpyxl.styles import colors

"""
import xlrd
import xlwt
from xlutils.copy import copy

filee = "d:/asd.xls"
wrokbook = xlrd.open_workbook(filee)

wb = copy(wrokbook)
sheet001 = wb.get_sheet(0)
sheet001.write(0,0,"NUMER")
wb.save("asd2.xls")


sheet = wrokbook.sheet_by_index(0)
print sheet.cell_value(0,0)


print sheet.nrows
print sheet.ncols

for xol in range(sheet.ncols):
    print sheet.cell_value(0,xol)
    print sheet.cell_type(0,xol)

print sheet.row_values(2)

print sheet.cell(0,0)

print wrokbook.nsheets
print wrokbook.sheet_names()

workbook2 = xlwt.Workbook(encoding="utf-8")
sheet1 = workbook2.add_sheet("P1")
sheet2 = workbook2.add_sheet("P2")
sheet3 = workbook2.add_sheet("P3")

czcionka = xlwt.Font()
czcionka.name = "Calibri"
czcionka.bold = True
czcionka.colour_index = 3


style1 =  xlwt.XFStyle()

style1.font = czcionka


sheet1.write(0,0,"bleble 1",style1)
sheet2.write(0,1,"bleble 2")
sheet3.write(1,0,xlwt.Formula("A1+A3"))



workbook2.save("Moj.xls")

"""

location = "asd2.xlsx"

book = load_workbook(location)

print book.get_sheet_names()

# get_active_sheet
sheet = book.active
# book.get_sheet_by_name(name)

print sheet.title


print sheet["A1"].value
print sheet["A1"].column
print sheet["A1"].row

print sheet.cell(row=1,column=2).value

# for

#Zapis 

ft = Font(color=colors.RED)
wb = Workbook()

ws = wb.active
ws.font = ft
ws.title = "TYTAASD"

ws["B2"] = 43
ws["B6"] = "=A5+A6"

cell = WriteOnlyCell(ws, value="hello world")
cell.font = Font(name='Courier', size=36)
cell.comment = Comment(text="A comment", author="Author's Name")

ws.append([cell, 3.14, 4])
cell2 = ws.cell(row=1,column=2).value = "DUPA"
cell2 = ws.cell(row=1,column=2).font =  Font(name='Calibri', size=55,color=colors.RED)
wb.save("asd2.xlsx")


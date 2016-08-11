'''
Created on 11 sie 2016

@author: Magda
'''
from openpyxl import Workbook
from openpyxl.writer.write_only import WriteOnlyCell
from openpyxl.comments import Comment
from openpyxl.styles import Font
from openpyxl import load_workbook
from openpyxl.styles import colors

location = "IntegrationTestsControl_CONN.xlsx"

wb = load_workbook(location)
ws = wb.active


def zad5():
    lista = []
    for x in range(1,86):
        for y in range (1,6):
            result = ws.cell(row=x,column=y).value
            lista.append(result)
    return lista

print zad5()
  
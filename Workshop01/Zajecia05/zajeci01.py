'''
Created on 11 sie 2016

@author: giberkrz
'''

from openpyxl import Workbook
from openpyxl.styles import *
from openpyxl.comments import Comment
from openpyxl.reader.excel import load_workbook

location = "IntegrationTestsControl_CONN.xlsx"

wb = load_workbook(location)

wb.get_sheet_names()

ws = wb.active

print ws.title

print ws["A1"].value

print ws.cell(row=1,column=1).value

for x in range(1,10):
    print ws.cell(row=x,column=2).value


filename = "nowy.xlsx"

wb = Workbook()

ws = wb.active

ws.title = "TUTULASDAD"

ws["B3"].value = "KKKK"


ws["A2"].value = "Krzysiek"
ws["A2"].font =  Font(name='Calibri', size=55,color=colors.RED)

ws["A2"].comment = Comment("komentarz","KG")

ws.append([1,2,3,4,5])
ws.append([1,2,3,4,5])

ws["B3"]="=SUM(A1,B1)"

wb.save(filename)


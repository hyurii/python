'''
Created on 12 sie 2016

@author: giberkrz
'''

from openpyxl import Workbook
from openpyxl.writer.write_only import WriteOnlyCell
from openpyxl.comments import Comment
from openpyxl.styles import Font
from openpyxl import load_workbook
from openpyxl.styles import colors


def tworzenie_listy(ws):
    xls_list = []
    xls_list2 = []
    for x in ws.rows:
        for y in x:
            xls_list2.append(y.value)
        xls_list.append(xls_list2)
        xls_list2 = []
    return xls_list


def liczba_not_tested(listunia):
    dziwna_lista = []
    liczba_nie_test = 0
    for x in listunia:
        if x[4] == "NOT TESTED":
            liczba_nie_test += 1
            dziwna_lista.append(x[2])
    return liczba_nie_test, dziwna_lista


def nowy_shieet(wb, ilosc, listunia, lokalizacja):
    ws = wb.create_sheet("cos_nowego")
    ws["A1"].value = "Liczba not Tested: %s" % ilosc
    for x in range(len(listunia)):
        ws.cell(row=x + 2, column=1).value = "%s" % listunia[x]
    wb.save(lokalizacja)


def colorowanki(wb, lokalizacja):
    ws = wb.get_sheet_by_name("IntegrationTestsControl_CONN")
    for x in ws.rows:
        if x[4].value == "NOT TESTED":
            x[4].font = Font(color=colors.RED)
    wb.save(lokalizacja)


if __name__ == '__main__':
    pass

location = "IntegrationTestsControl_CONN.xlsx"
wb = load_workbook(location)
ws = wb.active

colorowanki(wb, location)

#lista = tworzenie_listy(ws)
#losc, lista2 = liczba_not_tested(lista)

#nowy_shieet(wb, ilosc, lista2,location)



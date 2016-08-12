'''
Created on 12 sie 2016

@author: giberkrz
'''

from openpyxl import Workbook
from openpyxl.writer.write_only import WriteOnlyCell
from openpyxl.comments import Comment
from openpyxl.styles import Font
from openpyxl import load_workbook
from openpyxl.styles import colors


location = "IntegrationTestsControl_CONN.xlsx"
wb = load_workbook(location)

print wb.get_sheet_names()

ws = wb.active

print ws.title
print ws

print ws["A1"].value

print ws.cell(row=2,column=2).value

wb2 = Workbook(write_only=True)

ws = wb2.active

ws.title = "KKKK"

ws["A2"].value = "Krzysiek"

ws.cell(row=3,column=3).value = "=A1+A2"

ft = Font(color=colors.RED,size=55,bold=True,name="Calibri")
comment = Comment("hhh","KG")
ws['A2'].font = ft
ws['A2'].comment = comment
ws.append([1,2,3,4,45,4235423])

wb2.save("Z2.xlsx")
 





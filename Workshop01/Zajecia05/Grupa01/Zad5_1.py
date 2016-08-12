'''
Created on 11 sie 2016

@author: Magda
'''
from openpyxl import Workbook
from openpyxl.writer.write_only import WriteOnlyCell
from openpyxl.comments import Comment
from openpyxl.styles import Font
from openpyxl import load_workbook
from openpyxl.styles import colors
import list_parsing

location = "IntegrationTestsControl_CONN.xlsx"

wb = load_workbook(location)
ws = wb.active


def zad5():
    lista = []
    lista2 = []
    for y in range (1,6):
        lista2 = []
        for x in range(1,86):
            result = ws.cell(row=x,column=y).value
            lista2.append(result)
        lista.append(lista2)
    return lista


lista = zad5()
print lista
print list_parsing.ListParsing(lista)


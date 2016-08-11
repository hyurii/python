from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl import styles

location = "C:\Users\Tomek\Desktop\python_git\workspace\Workshop01\Zajecia05\IntegrationTestsControl_CONN.xlsx"

wb = load_workbook(location)

# print wb.get_sheet_names()

ws = wb.active

print ws.title

print ws["A1"].value

print ws.cell(row=1,column=1).value

for x in range(1,10):
    print ws.cell(row=x,column=1).value

'''
wb = Workbook()

#automatycznie tworzy sheet z domyslna nazwa
ws = wb.active
ws.title = "Tutorial"
ws["B3"] = "tutututu"

wb.save("new.xlsx")
'''

# ws["A2"].font = Font(name="Calibri", size = 55, color = colors.RED)
#ws["A2].comment = Comment("komentarz", "KG")


# ws["B3"]="A1+A2"